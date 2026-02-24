"""
LACO Membership Management System
Labour and Civic Organization - South Africa
Version 1.1 - Fixed save button + Reports window
"""

import tkinter as tk
from PIL import Image, ImageTk
from tkinter import ttk, messagebox, filedialog
import os
import platform
import subprocess
from datetime import datetime
from laco_database import (
    create_tables, add_member, update_member, delete_member,
    get_all_members, search_members, get_member_by_id, get_stats
)

# ============================================================================
# COLOUR THEME
# ============================================================================
COLORS = {
    "bg":        "#1a4a1a",   # dark green       - main background (logo bg)
    "panel":     "#122f12",   # deeper green     - panels / stats bar
    "card":      "#0d220d",   # very dark green  - cards
    "accent":    "#c8a400",   # gold/yellow      - highlights (logo text)
    "gold":      "#f0c800",   # bright gold      - headings / labels
    "green":     "#2d7a2d",   # mid green        - save / active
    "text":      "#f0c800",   # gold             - primary text on dark
    "subtext":   "#999999",   # mid grey         - secondary text
    "entry_bg":  "#0d220d",   # dark green       - input field bg
    "header_bg": "#0a1a0a",   # near-black green - header bar
    "red":       "#cc2200",   # red              - delete/danger (logo red)
    "black":     "#111111",   # near-black       - borders/circle
}

PROVINCES  = ["Gauteng","KwaZulu-Natal","Eastern Cape","Western Cape",
               "Northern Cape","North West","Limpopo","Mpumalanga","Free State"]
TITLES     = ["Mr","Mrs","Ms","Dr","Prof"]
CATEGORIES = ["Ordinary (R50/month)","Additional Contribution","Non-Member Donation"]
STATUSES   = ["Pending","Active","Inactive","Suspended"]


# ============================================================================
# HELPER WIDGETS
# ============================================================================

def styled_entry(parent, width=30, textvariable=None):
    return tk.Entry(parent, width=width, font=("Arial", 10),
                    bg=COLORS["entry_bg"], fg=COLORS["gold"],
                    insertbackground="black",
                    relief="flat", bd=4, textvariable=textvariable)

def styled_button(parent, text, command, color=None, width=15):
    bg = color or COLORS["accent"]
    # Darken active colour slightly for hover effect
    btn = tk.Button(parent, text=text, command=command,
                     font=("Arial", 10, "bold"), bg=bg, fg="#000000",
                     activebackground=bg, activeforeground="#000000",
                     relief="raised", bd=2,
                     padx=14, pady=7, cursor="hand2", width=width,
                     highlightbackground="#ffffff",
                     highlightthickness=1)
    return btn

def styled_combo(parent, values, textvariable=None, width=28):
    return ttk.Combobox(parent, values=values, textvariable=textvariable,
                        width=width, font=("Arial", 10), state="readonly")


# ============================================================================
# REPORTS WINDOW
# ============================================================================

class ReportsWindow(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("LACO - Reports & Statistics")
        self.geometry("720x600")
        self.configure(bg=COLORS["bg"])
        self.grab_set()
        self._build()
        self._load()

    def _build(self):
        # Header
        hdr = tk.Frame(self, bg=COLORS["header_bg"], pady=12)
        hdr.pack(fill="x")
        tk.Label(hdr, text="Reports & Statistics",
                 font=("Arial", 15, "bold"), fg="black",
                 bg=COLORS["header_bg"]).pack()

        # Summary cards
        cards_frame = tk.Frame(self, bg=COLORS["panel"], pady=14)
        cards_frame.pack(fill="x")
        self.stat_labels = {}
        for key, label, color in [
            ("total",           "Total Members",    COLORS["gold"]),
            ("active",          "Active",           COLORS["green"]),
            ("pending",         "Pending",          "#c8a400"),
            ("monthly_revenue", "Monthly Revenue",  COLORS["accent"]),
        ]:
            card = tk.Frame(cards_frame, bg=COLORS["card"], padx=20, pady=10)
            card.pack(side="left", expand=True, fill="x", padx=10, pady=4)
            tk.Label(card, text=label, font=("Arial", 9),
                     fg=COLORS["subtext"], bg=COLORS["card"]).pack()
            lbl = tk.Label(card, text="--", font=("Arial", 20, "bold"),
                           fg=color, bg=COLORS["card"])
            lbl.pack()
            self.stat_labels[key] = lbl

        # Tables area
        section_frame = tk.Frame(self, bg=COLORS["bg"])
        section_frame.pack(fill="both", expand=True, padx=15, pady=10)

        style = ttk.Style()
        style.configure("Report.Treeview",
                        background=COLORS["panel"], foreground=COLORS["gold"],
                        fieldbackground=COLORS["panel"], rowheight=24,
                        font=("Arial", 10))
        style.configure("Report.Treeview.Heading",
                        background=COLORS["card"], foreground=COLORS["accent"],
                        font=("Arial", 10, "bold"))

        # Province breakdown
        tk.Label(section_frame, text="  Members by Province",
                 font=("Arial", 11, "bold"), fg=COLORS["gold"],
                 bg=COLORS["card"], anchor="w", pady=6).pack(fill="x", pady=(0,4))

        prov_cols = ("Province", "Total", "Active", "Pending", "Monthly Revenue")
        self.prov_tree = ttk.Treeview(section_frame, columns=prov_cols,
                                      show="headings", height=9,
                                      style="Report.Treeview")
        for col in prov_cols:
            self.prov_tree.heading(col, text=col)
            self.prov_tree.column(col, width=120, anchor="center")
        self.prov_tree.column("Province", width=170, anchor="w")
        self.prov_tree.pack(fill="x")

        # Status breakdown
        tk.Label(section_frame, text="  Members by Status",
                 font=("Arial", 11, "bold"), fg=COLORS["gold"],
                 bg=COLORS["card"], anchor="w", pady=6).pack(fill="x", pady=(14,4))

        status_cols = ("Status", "Count", "% of Total")
        self.status_tree = ttk.Treeview(section_frame, columns=status_cols,
                                        show="headings", height=5,
                                        style="Report.Treeview")
        for col in status_cols:
            self.status_tree.heading(col, text=col)
            self.status_tree.column(col, width=180, anchor="center")
        self.status_tree.pack(fill="x")

        # Buttons
        btn_bar = tk.Frame(self, bg=COLORS["bg"], pady=12)
        btn_bar.pack(fill="x", padx=15)
        styled_button(btn_bar, "Refresh", self._load,
                      color="#c8a400", width=12).pack(side="left", padx=4)
        styled_button(btn_bar, "Print PDF Report", self._print_pdf,
                      color="#2d7a2d", width=18).pack(side="left", padx=4)
        styled_button(btn_bar, "Print Member List", self._print_member_list,
                      color="#2d7a2d", width=18).pack(side="left", padx=4)
        styled_button(btn_bar, "Export Excel", self._export_excel,
                      color="#1a5276", width=15).pack(side="left", padx=4)
        styled_button(btn_bar, "Close", self.destroy,
                      color="#cc2200", width=12).pack(side="left", padx=4)

    def _load(self):
        s = get_stats()
        self.stat_labels["total"].config(text=str(s["total"]))
        self.stat_labels["active"].config(text=str(s["active"]))
        self.stat_labels["pending"].config(text=str(s["pending"]))
        self.stat_labels["monthly_revenue"].config(text=f"R {s['monthly_revenue']:,.0f}")

        from laco_database import get_conn
        conn = get_conn()
        c = conn.cursor()

        for item in self.prov_tree.get_children():
            self.prov_tree.delete(item)
        c.execute("""
            SELECT province,
                   COUNT(*) as total,
                   SUM(CASE WHEN status='Active' THEN 1 ELSE 0 END),
                   SUM(CASE WHEN status='Pending' THEN 1 ELSE 0 END),
                   SUM(CASE WHEN status='Active' THEN monthly_amount ELSE 0 END)
            FROM members GROUP BY province ORDER BY total DESC
        """)
        for row in c.fetchall():
            self.prov_tree.insert("", "end", values=(
                row[0] or "Unknown", row[1], row[2], row[3], f"R {row[4]:,.0f}"
            ))

        for item in self.status_tree.get_children():
            self.status_tree.delete(item)
        total = max(s["total"], 1)
        c.execute("SELECT status, COUNT(*) FROM members GROUP BY status ORDER BY 2 DESC")
        for row in c.fetchall():
            self.status_tree.insert("", "end", values=(
                row[0], row[1], f"{row[1]/total*100:.1f}%"
            ))
        conn.close()

    def _print_member_list(self):
        """Generate a clean printable PDF of all members with full details."""
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib import colors as rl_colors
            from reportlab.lib.styles import ParagraphStyle
            from reportlab.lib.units import mm
            from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                            Paragraph, Spacer, HRFlowable)
            from reportlab.lib.enums import TA_CENTER, TA_LEFT
        except ImportError:
            messagebox.showerror("Missing Library",
                "reportlab is required. Run: pip3 install reportlab", parent=self)
            return

        import os, platform, subprocess
        from laco_database import get_conn
        from tkinter import filedialog

        filepath = filedialog.asksaveasfilename(
            parent=self,
            defaultextension=".pdf",
            filetypes=[("PDF", "*.pdf")],
            initialfile=f"LACO_MemberList_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf",
            title="Save Member List PDF"
        )
        if not filepath:
            return

        try:
            # Landscape A4 for more column space
            doc = SimpleDocTemplate(
                filepath, pagesize=landscape(A4),
                rightMargin=10*mm, leftMargin=10*mm,
                topMargin=10*mm, bottomMargin=10*mm
            )

            DARK_GREEN = rl_colors.HexColor("#0D220D")
            MID_GREEN  = rl_colors.HexColor("#1A4A1A")
            GOLD       = rl_colors.HexColor("#F0C800")
            LIGHT_GOLD = rl_colors.HexColor("#C8A400")
            WHITE      = rl_colors.white
            STATUS_C   = {
                "Active":    rl_colors.HexColor("#4dcc4d"),
                "Pending":   rl_colors.HexColor("#f0c800"),
                "Inactive":  rl_colors.HexColor("#999999"),
                "Suspended": rl_colors.HexColor("#ff4422"),
            }

            title_style = ParagraphStyle("t", fontName="Helvetica-Bold",
                fontSize=14, textColor=GOLD, backColor=DARK_GREEN,
                alignment=TA_CENTER, spaceAfter=2, borderPadding=8)
            sub_style = ParagraphStyle("s", fontName="Helvetica",
                fontSize=8, textColor=rl_colors.HexColor("#999999"),
                alignment=TA_CENTER, spaceAfter=4)
            sec_style = ParagraphStyle("sec", fontName="Helvetica-Bold",
                fontSize=10, textColor=GOLD, backColor=MID_GREEN,
                spaceBefore=8, spaceAfter=3, borderPadding=4)

            def make_table(data, col_widths, status_col=None):
                t = Table(data, colWidths=col_widths, repeatRows=1)
                ts = [
                    ("BACKGROUND",    (0,0), (-1,0), DARK_GREEN),
                    ("TEXTCOLOR",     (0,0), (-1,0), GOLD),
                    ("FONTNAME",      (0,0), (-1,0), "Helvetica-Bold"),
                    ("FONTSIZE",      (0,0), (-1,-1), 7),
                    ("ALIGN",         (0,0), (-1,-1), "LEFT"),
                    ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
                    ("BOX",           (0,0), (-1,-1), 1,   GOLD),
                    ("GRID",          (0,0), (-1,-1), 0.3, LIGHT_GOLD),
                    ("TOPPADDING",    (0,0), (-1,-1), 3),
                    ("BOTTOMPADDING", (0,0), (-1,-1), 3),
                    ("LEFTPADDING",   (0,0), (-1,-1), 4),
                ]
                for i in range(1, len(data)):
                    bg = MID_GREEN if i % 2 == 0 else DARK_GREEN
                    ts.append(("BACKGROUND", (0,i), (-1,i), bg))
                    ts.append(("TEXTCOLOR",  (0,i), (-1,i), WHITE))
                    if status_col is not None:
                        sv = data[i][status_col]
                        sc = STATUS_C.get(sv, WHITE)
                        ts.append(("TEXTCOLOR",  (status_col,i), (status_col,i), sc))
                        ts.append(("FONTNAME",   (status_col,i), (status_col,i), "Helvetica-Bold"))
                t.setStyle(TableStyle(ts))
                return t

            # Fetch all members
            conn = get_conn()
            c = conn.cursor()
            c.execute("""
                SELECT membership_number, title, surname, firstnames, initials,
                       id_number, date_of_birth, gender,
                       cell_number, tel_number, email,
                       postal_address, postal_code, province, municipality, ward,
                       membership_category, monthly_amount, status, date_joined,
                       bank_account_holder, bank_name, account_number,
                       branch_name, branch_code, debit_order_date
                FROM members ORDER BY surname ASC
            """)
            members = c.fetchall()
            conn.close()

            total = len(members)
            story = []

            # ── Cover header ─────────────────────────────────────────────
            story.append(Paragraph("LABOUR AND CIVIC ORGANIZATION — MEMBER LIST", title_style))
            story.append(Paragraph(
                f"Printed: {datetime.now().strftime('%d %B %Y  %H:%M')}  |  "
                f"Total Members: {total}  |  lacosouthafrica@outlook.com  |  071 526 9250",
                sub_style))
            story.append(HRFlowable(width="100%", thickness=2, color=GOLD))
            story.append(Spacer(1, 4*mm))

            if not members:
                story.append(Paragraph("No members found in the database.", sub_style))
            else:
                # ── TABLE 1: Personal & Contact ───────────────────────────
                story.append(Paragraph("Personal & Contact Details", sec_style))
                pc = [["#", "Mem No", "Title", "Surname", "First Names",
                        "ID Number", "DOB", "Gender", "Cell", "Tel", "Email",
                        "Status", "Joined"]]
                for idx, m in enumerate(members, 1):
                    pc.append([
                        str(idx), m[0] or "", m[1] or "",
                        m[2] or "", m[3] or "",
                        m[5] or "", str(m[6] or "")[:10], m[7] or "",
                        m[8] or "", m[9] or "", m[10] or "",
                        m[18] or "", str(m[19] or "")[:10]
                    ])
                story.append(make_table(pc,
                    [7*mm, 22*mm, 8*mm, 22*mm, 26*mm,
                     26*mm, 18*mm, 10*mm, 20*mm, 20*mm,
                     36*mm, 14*mm, 18*mm],
                    status_col=11))
                story.append(Spacer(1, 5*mm))

                # ── TABLE 2: Address & Membership ─────────────────────────
                story.append(Paragraph("Address & Membership Details", sec_style))
                am = [["#", "Mem No", "Surname", "Postal Address", "Code",
                        "Province", "Municipality", "Ward", "Category", "Monthly (R)"]]
                for idx, m in enumerate(members, 1):
                    am.append([
                        str(idx), m[0] or "", m[2] or "",
                        m[11] or "", m[12] or "",
                        m[13] or "", m[14] or "", m[15] or "",
                        m[16] or "", f"R {m[17] or 0:,.0f}"
                    ])
                story.append(make_table(am,
                    [7*mm, 22*mm, 25*mm, 45*mm, 12*mm,
                     22*mm, 28*mm, 12*mm, 38*mm, 18*mm]))
                story.append(Spacer(1, 5*mm))

                # ── TABLE 3: Banking Details ──────────────────────────────
                story.append(Paragraph("Banking & Debit Order Details", sec_style))
                bk = [["#", "Mem No", "Surname", "Account Holder",
                        "Bank", "Account No", "Branch Name", "Branch Code", "Debit Date"]]
                for idx, m in enumerate(members, 1):
                    bk.append([
                        str(idx), m[0] or "", m[2] or "",
                        m[20] or "", m[21] or "", m[22] or "",
                        m[23] or "", m[24] or "", m[25] or ""
                    ])
                story.append(make_table(bk,
                    [7*mm, 22*mm, 25*mm, 30*mm,
                     25*mm, 28*mm, 25*mm, 18*mm, 16*mm]))

            # ── Footer ───────────────────────────────────────────────────
            story.append(Spacer(1, 5*mm))
            story.append(HRFlowable(width="100%", thickness=1, color=GOLD))
            story.append(Paragraph(
                "Labour and Civic Organization (LACO)  |  403 Delta Towers, "
                "300 Anton Lembede Street, Durban 4000  |  lacosouthafrica@outlook.com",
                sub_style))

            doc.build(story)

            messagebox.showinfo("Member List Ready", f"Saved to: {filepath}. Opening for printing...", parent=self)

            if platform.system() == "Darwin":
                subprocess.call(["open", filepath])
            elif platform.system() == "Windows":
                os.startfile(filepath)
            else:
                subprocess.call(["xdg-open", filepath])

        except Exception as e:
            messagebox.showerror("Print Error",
                f"Could not generate member list: {str(e)}", parent=self)

    def _export_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import (Font, PatternFill, Alignment,
                                         Border, Side)
            from openpyxl.utils import get_column_letter
        except ImportError:
            messagebox.showerror("Missing Library",
                "openpyxl is required. Run: pip3 install openpyxl", parent=self)
            return

        import os, platform, subprocess
        from laco_database import get_conn, get_stats
        from tkinter import filedialog

        filepath = filedialog.asksaveasfilename(
            parent=self,
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=f"LACO_Report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            title="Save Excel Report"
        )
        if not filepath:
            return

        try:
            wb = openpyxl.Workbook()

            # ── Colour definitions ────────────────────────────────────────
            DARK_GREEN  = "0D220D"
            MID_GREEN   = "1A4A1A"
            GOLD        = "F0C800"
            LIGHT_GOLD  = "C8A400"
            RED         = "CC2200"
            WHITE       = "FFFFFF"
            GREY        = "999999"
            ALT_ROW     = "122F12"

            def hdr_font(size=11, color=GOLD):
                return Font(name="Arial", bold=True, size=size, color=color)

            def cell_font(size=10, color=WHITE, bold=False):
                return Font(name="Arial", size=size, color=color, bold=bold)

            def fill(hex_color):
                return PatternFill("solid", fgColor=hex_color)

            def center():
                return Alignment(horizontal="center", vertical="center", wrap_text=True)

            def left():
                return Alignment(horizontal="left", vertical="center", wrap_text=True)

            def thin_border():
                s = Side(style="thin", color=LIGHT_GOLD)
                return Border(left=s, right=s, top=s, bottom=s)

            def style_header_row(ws, row, cols, bg=DARK_GREEN, fg=GOLD, height=20):
                ws.row_dimensions[row].height = height
                for col in range(1, cols+1):
                    c = ws.cell(row=row, column=col)
                    c.font = hdr_font(color=fg)
                    c.fill = fill(bg)
                    c.alignment = center()
                    c.border = thin_border()

            def style_data_row(ws, row, cols, alt=False):
                bg = ALT_ROW if alt else DARK_GREEN
                for col in range(1, cols+1):
                    c = ws.cell(row=row, column=col)
                    c.font = cell_font()
                    c.fill = fill(bg)
                    c.alignment = center()
                    c.border = thin_border()

            s = get_stats()
            conn = get_conn()
            c = conn.cursor()

            # ══════════════════════════════════════════════════════════════
            # SHEET 1 - Summary
            # ══════════════════════════════════════════════════════════════
            ws1 = wb.active
            ws1.title = "Summary"
            ws1.sheet_view.showGridLines = False
            ws1.column_dimensions["A"].width = 28
            ws1.column_dimensions["B"].width = 20

            # Title
            ws1.merge_cells("A1:B1")
            t = ws1["A1"]
            t.value = "LABOUR AND CIVIC ORGANIZATION"
            t.font = hdr_font(size=14, color=GOLD)
            t.fill = fill(DARK_GREEN)
            t.alignment = center()
            ws1.row_dimensions[1].height = 28

            ws1.merge_cells("A2:B2")
            t2 = ws1["A2"]
            t2.value = f"Membership Report — {datetime.now().strftime('%d %B %Y %H:%M')}"
            t2.font = cell_font(size=9, color=GREY)
            t2.fill = fill(DARK_GREEN)
            t2.alignment = center()
            ws1.row_dimensions[2].height = 16

            # Stats
            headers = ["Metric", "Value"]
            ws1.append(headers)
            style_header_row(ws1, 3, 2)

            rows = [
                ("Total Members", s["total"]),
                ("Active Members", s["active"]),
                ("Pending Members", s["pending"]),
                ("Inactive/Suspended", s["total"] - s["active"] - s["pending"]),
                ("Monthly Revenue", f"R {s['monthly_revenue']:,.0f}"),
            ]
            for i, (metric, val) in enumerate(rows, 4):
                ws1.cell(row=i, column=1, value=metric)
                ws1.cell(row=i, column=2, value=str(val))
                style_data_row(ws1, i, 2, alt=(i % 2 == 0))
                ws1.cell(row=i, column=1).alignment = left()

            # ══════════════════════════════════════════════════════════════
            # SHEET 2 - Province Breakdown
            # ══════════════════════════════════════════════════════════════
            ws2 = wb.create_sheet("By Province")
            ws2.sheet_view.showGridLines = False
            for col, width in zip("ABCDEF", [28, 12, 12, 12, 12, 18]):
                ws2.column_dimensions[col].width = width

            ws2.merge_cells("A1:F1")
            t = ws2["A1"]
            t.value = "Members by Province"
            t.font = hdr_font(size=13, color=GOLD)
            t.fill = fill(DARK_GREEN)
            t.alignment = center()
            ws2.row_dimensions[1].height = 24

            prov_headers = ["Province", "Total", "Active", "Pending", "Inactive", "Revenue (R)"]
            ws2.append(prov_headers)
            style_header_row(ws2, 2, 6)

            c.execute("""
                SELECT province,
                       COUNT(*),
                       SUM(CASE WHEN status='Active' THEN 1 ELSE 0 END),
                       SUM(CASE WHEN status='Pending' THEN 1 ELSE 0 END),
                       SUM(CASE WHEN status='Inactive' THEN 1 ELSE 0 END),
                       SUM(CASE WHEN status='Active' THEN monthly_amount ELSE 0 END)
                FROM members GROUP BY province ORDER BY 2 DESC
            """)
            for i, row in enumerate(c.fetchall(), 3):
                ws2.append([row[0] or "Unknown", row[1], row[2] or 0,
                            row[3] or 0, row[4] or 0, row[5] or 0])
                style_data_row(ws2, i, 6, alt=(i % 2 == 0))
                ws2.cell(row=i, column=1).alignment = left()
                ws2.cell(row=i, column=6).number_format = 'R #,##0.00'

            # ══════════════════════════════════════════════════════════════
            # SHEET 3 - Status Breakdown
            # ══════════════════════════════════════════════════════════════
            ws3 = wb.create_sheet("By Status")
            ws3.sheet_view.showGridLines = False
            for col, width in zip("ABC", [20, 14, 16]):
                ws3.column_dimensions[col].width = width

            ws3.merge_cells("A1:C1")
            t = ws3["A1"]
            t.value = "Members by Status"
            t.font = hdr_font(size=13, color=GOLD)
            t.fill = fill(DARK_GREEN)
            t.alignment = center()
            ws3.row_dimensions[1].height = 24

            ws3.append(["Status", "Count", "% of Total"])
            style_header_row(ws3, 2, 3)

            total = max(s["total"], 1)
            c.execute("SELECT status, COUNT(*) FROM members GROUP BY status ORDER BY 2 DESC")
            status_colors_map = {
                "Active": "2D7A2D", "Pending": "C8A400",
                "Inactive": "555555", "Suspended": "CC2200"
            }
            for i, row in enumerate(c.fetchall(), 3):
                pct = f"{row[1]/total*100:.1f}%"
                ws3.append([row[0], row[1], pct])
                style_data_row(ws3, i, 3, alt=(i % 2 == 0))
                sc = status_colors_map.get(row[0], WHITE)
                ws3.cell(row=i, column=1).font = Font(name="Arial", size=10,
                                                      color=sc, bold=True)

            # ══════════════════════════════════════════════════════════════
            # Fetch all member data once for sheets 4, 5, 6
            # ══════════════════════════════════════════════════════════════
            c.execute("""
                SELECT membership_number, title, surname, firstnames, initials,
                       id_number, date_of_birth, gender,
                       cell_number, tel_number, email,
                       postal_address, postal_code, province, municipality, ward,
                       membership_category, monthly_amount, status, date_joined,
                       bank_account_holder, bank_name, account_number,
                       branch_name, branch_code, debit_order_date
                FROM members ORDER BY surname ASC
            """)
            all_members = c.fetchall()
            conn.close()

            # ── Helper: build a sheet with title + headers + data ─────────
            def build_sheet(title_text, headers, rows, col_widths_list):
                ws = wb.create_sheet(title_text)
                ws.sheet_view.showGridLines = False
                ncols = len(headers)
                for ci, w in enumerate(col_widths_list, 1):
                    ws.column_dimensions[get_column_letter(ci)].width = w
                last_col = get_column_letter(ncols)
                ws.merge_cells(f"A1:{last_col}1")
                t = ws["A1"]
                t.value = title_text
                t.font = hdr_font(size=13, color=GOLD)
                t.fill = fill(DARK_GREEN)
                t.alignment = center()
                ws.row_dimensions[1].height = 24
                ws.append(headers)
                style_header_row(ws, 2, ncols, height=20)
                for ri, row_data in enumerate(rows, 3):
                    ws.append(row_data)
                    style_data_row(ws, ri, ncols, alt=(ri % 2 == 0))
                    for ci in range(1, ncols+1):
                        ws.cell(row=ri, column=ci).alignment = left()
                if not rows:
                    ws.append(["No data"] + ["-"]*(ncols-1))
                    style_data_row(ws, 3, ncols)
                return ws

            # ══════════════════════════════════════════════════════════════
            # SHEET 4 — Personal & Contact
            # ══════════════════════════════════════════════════════════════
            pc_headers = ["Mem No","Title","Surname","First Names","Initials",
                          "ID Number","Date of Birth","Gender",
                          "Cell","Tel","Email","Status","Date Joined"]
            pc_rows = []
            for m in all_members:
                pc_rows.append([
                    m[0] or "", m[1] or "", m[2] or "", m[3] or "", m[4] or "",
                    m[5] or "", str(m[6] or "")[:10], m[7] or "",
                    m[8] or "", m[9] or "", m[10] or "",
                    m[18] or "", str(m[19] or "")[:10]
                ])
            ws4 = build_sheet("Personal & Contact", pc_headers, pc_rows,
                              [18,8,20,22,10,18,14,10,16,16,28,12,14])
            # Colour status column (col 12)
            for ri, m in enumerate(all_members, 3):
                sv = m[18] or ""
                sc = status_colors_map.get(sv, WHITE)
                ws4.cell(row=ri, column=12).font = Font(name="Arial", size=10,
                                                        color=sc, bold=True)

            # ══════════════════════════════════════════════════════════════
            # SHEET 5 — Address & Membership
            # ══════════════════════════════════════════════════════════════
            am_headers = ["Mem No","Surname","Postal Address","Postal Code",
                          "Province","Municipality","Ward",
                          "Membership Category","Monthly Amt (R)","Status"]
            am_rows = []
            for m in all_members:
                am_rows.append([
                    m[0] or "", m[2] or "", m[11] or "", m[12] or "",
                    m[13] or "", m[14] or "", m[15] or "",
                    m[16] or "", m[17] or 0, m[18] or ""
                ])
            ws5 = build_sheet("Address & Membership", am_headers, am_rows,
                              [18,22,35,12,18,22,10,28,16,12])
            for ri, m in enumerate(all_members, 3):
                ws5.cell(row=ri, column=9).number_format = 'R #,##0.00'
                sv = m[18] or ""
                sc = status_colors_map.get(sv, WHITE)
                ws5.cell(row=ri, column=10).font = Font(name="Arial", size=10,
                                                        color=sc, bold=True)

            # ══════════════════════════════════════════════════════════════
            # SHEET 6 — Banking Details
            # ══════════════════════════════════════════════════════════════
            bk_headers = ["Mem No","Surname","First Names",
                          "Account Holder","Bank Name",
                          "Account Number","Branch Name","Branch Code","Debit Date"]
            bk_rows = []
            for m in all_members:
                bk_rows.append([
                    m[0] or "", m[2] or "", m[3] or "",
                    m[20] or "", m[21] or "",
                    m[22] or "", m[23] or "", m[24] or "", m[25] or ""
                ])
            build_sheet("Banking Details", bk_headers, bk_rows,
                        [18,22,22,26,22,24,22,16,14])
            wb.save(filepath)

            messagebox.showinfo("Excel Ready",
                f"Report exported to: {filepath}", parent=self)

            if platform.system() == "Darwin":
                subprocess.call(["open", filepath])
            elif platform.system() == "Windows":
                os.startfile(filepath)
            else:
                subprocess.call(["xdg-open", filepath])

        except Exception as e:
            messagebox.showerror("Excel Error", f"Could not export Excel: {str(e)}", parent=self)

    def _print_pdf(self):
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors as rl_colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import mm
            from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                            Paragraph, Spacer, HRFlowable)
            from reportlab.lib.enums import TA_CENTER, TA_LEFT
        except ImportError:
            messagebox.showerror("Missing Library",
                "reportlab is required. Run: pip3 install reportlab", parent=self)
            return

        import tempfile, os, platform, subprocess
        from laco_database import get_conn, get_stats

        # Ask where to save
        from tkinter import filedialog
        filepath = filedialog.asksaveasfilename(
            parent=self,
            defaultextension=".pdf",
            filetypes=[("PDF", "*.pdf")],
            initialfile=f"LACO_Report_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf",
            title="Save PDF Report"
        )
        if not filepath:
            return

        try:
            doc = SimpleDocTemplate(
                filepath, pagesize=A4,
                rightMargin=15*mm, leftMargin=15*mm,
                topMargin=15*mm, bottomMargin=15*mm
            )

            # Colour palette matching logo
            DARK_GREEN  = rl_colors.HexColor("#0d220d")
            MID_GREEN   = rl_colors.HexColor("#1a4a1a")
            GOLD        = rl_colors.HexColor("#f0c800")
            RED         = rl_colors.HexColor("#cc2200")
            LIGHT_GOLD  = rl_colors.HexColor("#c8a400")
            WHITE       = rl_colors.white
            GREY        = rl_colors.HexColor("#999999")

            styles = getSampleStyleSheet()
            title_style = ParagraphStyle("title",
                fontName="Helvetica-Bold", fontSize=18,
                textColor=GOLD, backColor=DARK_GREEN,
                alignment=TA_CENTER, spaceAfter=4, spaceBefore=4,
                leftIndent=0, borderPadding=10)
            sub_style = ParagraphStyle("sub",
                fontName="Helvetica", fontSize=10,
                textColor=GREY, alignment=TA_CENTER, spaceAfter=2)
            section_style = ParagraphStyle("section",
                fontName="Helvetica-Bold", fontSize=12,
                textColor=GOLD, backColor=DARK_GREEN,
                spaceBefore=12, spaceAfter=6,
                leftIndent=4, borderPadding=6)
            sub_section_style = ParagraphStyle("subsection",
                fontName="Helvetica-Bold", fontSize=10,
                textColor=LIGHT_GOLD, backColor=MID_GREEN,
                spaceBefore=8, spaceAfter=4,
                leftIndent=4, borderPadding=4)

            story = []

            # ── Title block ──────────────────────────────────────────────
            story.append(Paragraph("LABOUR AND CIVIC ORGANIZATION", title_style))
            story.append(Paragraph("Membership Report", title_style))
            story.append(Paragraph(
                f"Generated: {datetime.now().strftime('%d %B %Y  %H:%M')} "
                f"| Address: 403 Delta Towers, 300 Anton Lembede Street, Durban 4000",
                sub_style))
            story.append(HRFlowable(width="100%", thickness=2, color=GOLD))
            story.append(Spacer(1, 8*mm))

            # ── Summary Stats ────────────────────────────────────────────
            s = get_stats()
            story.append(Paragraph("Summary Statistics", section_style))
            summary_data = [
                ["Total Members", "Active", "Pending", "Monthly Revenue"],
                [str(s["total"]), str(s["active"]), str(s["pending"]),
                 f"R {s['monthly_revenue']:,.0f}"],
            ]
            summary_table = Table(summary_data, colWidths=[45*mm, 45*mm, 45*mm, 45*mm])
            summary_table.setStyle(TableStyle([
                ("BACKGROUND",  (0,0), (-1,0), DARK_GREEN),
                ("TEXTCOLOR",   (0,0), (-1,0), GOLD),
                ("FONTNAME",    (0,0), (-1,0), "Helvetica-Bold"),
                ("FONTSIZE",    (0,0), (-1,0), 11),
                ("ALIGN",       (0,0), (-1,-1), "CENTER"),
                ("BACKGROUND",  (0,1), (-1,1), MID_GREEN),
                ("TEXTCOLOR",   (0,1), (-1,1), WHITE),
                ("FONTNAME",    (0,1), (-1,1), "Helvetica-Bold"),
                ("FONTSIZE",    (0,1), (-1,1), 14),
                ("ROWBACKGROUNDS", (0,1), (-1,-1), [MID_GREEN]),
                ("BOX",         (0,0), (-1,-1), 1.5, GOLD),
                ("GRID",        (0,0), (-1,-1), 0.5, LIGHT_GOLD),
                ("TOPPADDING",  (0,0), (-1,-1), 8),
                ("BOTTOMPADDING",(0,0),(-1,-1), 8),
            ]))
            story.append(summary_table)
            story.append(Spacer(1, 6*mm))

            # ── Province Breakdown ───────────────────────────────────────
            conn = get_conn()
            c = conn.cursor()

            story.append(Paragraph("Members by Province", section_style))
            c.execute("""
                SELECT province,
                       COUNT(*) as total,
                       SUM(CASE WHEN status='Active' THEN 1 ELSE 0 END),
                       SUM(CASE WHEN status='Pending' THEN 1 ELSE 0 END),
                       SUM(CASE WHEN status='Inactive' THEN 1 ELSE 0 END),
                       SUM(CASE WHEN status='Active' THEN monthly_amount ELSE 0 END)
                FROM members GROUP BY province ORDER BY total DESC
            """)
            prov_rows = c.fetchall()

            prov_data = [["Province", "Total", "Active", "Pending", "Inactive", "Revenue"]]
            for row in prov_rows:
                prov_data.append([
                    row[0] or "Unknown", str(row[1]), str(row[2] or 0),
                    str(row[3] or 0), str(row[4] or 0), f"R {row[5]:,.0f}"
                ])
            if len(prov_data) == 1:
                prov_data.append(["No data", "-", "-", "-", "-", "-"])

            prov_table = Table(prov_data, colWidths=[50*mm, 22*mm, 22*mm, 22*mm, 22*mm, 32*mm])
            prov_style = [
                ("BACKGROUND",   (0,0), (-1,0), DARK_GREEN),
                ("TEXTCOLOR",    (0,0), (-1,0), GOLD),
                ("FONTNAME",     (0,0), (-1,0), "Helvetica-Bold"),
                ("FONTSIZE",     (0,0), (-1,-1), 9),
                ("ALIGN",        (0,0), (-1,-1), "CENTER"),
                ("ALIGN",        (0,0), (0,-1),  "LEFT"),
                ("BOX",          (0,0), (-1,-1), 1, GOLD),
                ("GRID",         (0,0), (-1,-1), 0.3, LIGHT_GOLD),
                ("TOPPADDING",   (0,0), (-1,-1), 5),
                ("BOTTOMPADDING",(0,0), (-1,-1), 5),
            ]
            for i in range(1, len(prov_data)):
                bg = MID_GREEN if i % 2 == 0 else DARK_GREEN
                prov_style.append(("BACKGROUND", (0,i), (-1,i), bg))
                prov_style.append(("TEXTCOLOR",  (0,i), (-1,i), WHITE))
            prov_table.setStyle(TableStyle(prov_style))
            story.append(prov_table)
            story.append(Spacer(1, 6*mm))

            # ── Status Breakdown ─────────────────────────────────────────
            story.append(Paragraph("Members by Status", section_style))
            c.execute("SELECT status, COUNT(*) FROM members GROUP BY status ORDER BY 2 DESC")
            total = max(s["total"], 1)
            status_data = [["Status", "Count", "% of Total"]]
            status_colors = {"Active": "#2d7a2d", "Pending": "#c8a400",
                             "Inactive": "#555555", "Suspended": "#cc2200"}
            for row in c.fetchall():
                status_data.append([row[0], str(row[1]), f"{row[1]/total*100:.1f}%"])
            if len(status_data) == 1:
                status_data.append(["No data", "-", "-"])

            status_table = Table(status_data, colWidths=[60*mm, 60*mm, 60*mm])
            st_style = [
                ("BACKGROUND",   (0,0), (-1,0), DARK_GREEN),
                ("TEXTCOLOR",    (0,0), (-1,0), GOLD),
                ("FONTNAME",     (0,0), (-1,0), "Helvetica-Bold"),
                ("FONTSIZE",     (0,0), (-1,-1), 10),
                ("ALIGN",        (0,0), (-1,-1), "CENTER"),
                ("BOX",          (0,0), (-1,-1), 1, GOLD),
                ("GRID",         (0,0), (-1,-1), 0.3, LIGHT_GOLD),
                ("TOPPADDING",   (0,0), (-1,-1), 6),
                ("BOTTOMPADDING",(0,0), (-1,-1), 6),
            ]
            for i in range(1, len(status_data)):
                bg = MID_GREEN if i % 2 == 0 else DARK_GREEN
                st_style.append(("BACKGROUND", (0,i), (-1,i), bg))
                st_style.append(("TEXTCOLOR",  (0,i), (-1,i), WHITE))
            status_table.setStyle(TableStyle(st_style))
            story.append(status_table)
            story.append(Spacer(1, 6*mm))

            # ── Full Member List — Page break before for clean layout ────
            from reportlab.platypus import PageBreak
            story.append(PageBreak())
            story.append(Paragraph("Full Member List", section_style))

            c.execute("""
                SELECT membership_number, title, surname, firstnames, initials,
                       id_number, date_of_birth, gender,
                       cell_number, tel_number, email,
                       postal_address, postal_code, province, municipality, ward,
                       membership_category, monthly_amount, status, date_joined,
                       bank_account_holder, bank_name, account_number,
                       branch_name, branch_code, debit_order_date
                FROM members ORDER BY surname ASC
            """)
            members = c.fetchall()
            conn.close()

            # ── Section 1: Personal & Contact ────────────────────────────
            story.append(Paragraph("Personal & Contact Details", sub_section_style))
            pc_headers = ["Mem No", "Title", "Surname", "First Names",
                          "ID Number", "DOB", "Gender", "Cell", "Tel", "Email", "Status", "Joined"]
            pc_data = [pc_headers]
            for m in members:
                pc_data.append([
                    m[0] or "", m[1] or "", m[2] or "", m[3] or "",
                    m[5] or "", str(m[6] or "")[:10], m[7] or "",
                    m[8] or "", m[9] or "", m[10] or "",
                    m[18] or "", str(m[19] or "")[:10]
                ])
            if len(pc_data) == 1:
                pc_data.append(["No members"] + ["-"] * 11)

            pc_col_w = [22*mm, 10*mm, 22*mm, 28*mm, 26*mm, 18*mm,
                        12*mm, 22*mm, 22*mm, 32*mm, 16*mm, 18*mm]
            pc_table = Table(pc_data, colWidths=pc_col_w, repeatRows=1)
            pc_style_list = [
                ("BACKGROUND",    (0,0), (-1,0), DARK_GREEN),
                ("TEXTCOLOR",     (0,0), (-1,0), GOLD),
                ("FONTNAME",      (0,0), (-1,0), "Helvetica-Bold"),
                ("FONTSIZE",      (0,0), (-1,-1), 7),
                ("ALIGN",         (0,0), (-1,-1), "LEFT"),
                ("BOX",           (0,0), (-1,-1), 1, GOLD),
                ("GRID",          (0,0), (-1,-1), 0.3, LIGHT_GOLD),
                ("TOPPADDING",    (0,0), (-1,-1), 3),
                ("BOTTOMPADDING", (0,0), (-1,-1), 3),
            ]
            for i in range(1, len(pc_data)):
                bg = MID_GREEN if i % 2 == 0 else DARK_GREEN
                pc_style_list.append(("BACKGROUND", (0,i), (-1,i), bg))
                pc_style_list.append(("TEXTCOLOR",  (0,i), (-1,i), WHITE))
                status_val = pc_data[i][10]
                sc = rl_colors.HexColor(status_colors.get(status_val, "#ffffff"))
                pc_style_list.append(("TEXTCOLOR", (10,i), (10,i), sc))
                pc_style_list.append(("FONTNAME",  (10,i), (10,i), "Helvetica-Bold"))
            pc_table.setStyle(TableStyle(pc_style_list))
            story.append(pc_table)
            story.append(Spacer(1, 5*mm))

            # ── Section 2: Address & Membership ──────────────────────────
            story.append(Paragraph("Address & Membership Details", sub_section_style))
            am_headers = ["Mem No", "Surname", "Postal Address", "Code",
                          "Province", "Municipality", "Ward", "Category", "Amt (R)"]
            am_data = [am_headers]
            for m in members:
                am_data.append([
                    m[0] or "", m[2] or "", m[11] or "", m[12] or "",
                    m[13] or "", m[14] or "", m[15] or "",
                    m[16] or "", f"R{m[17] or 0:,.0f}"
                ])
            if len(am_data) == 1:
                am_data.append(["No members"] + ["-"] * 8)

            am_col_w = [22*mm, 25*mm, 40*mm, 14*mm, 22*mm, 26*mm, 12*mm, 28*mm, 18*mm]
            am_table = Table(am_data, colWidths=am_col_w, repeatRows=1)
            am_style_list = [
                ("BACKGROUND",    (0,0), (-1,0), DARK_GREEN),
                ("TEXTCOLOR",     (0,0), (-1,0), GOLD),
                ("FONTNAME",      (0,0), (-1,0), "Helvetica-Bold"),
                ("FONTSIZE",      (0,0), (-1,-1), 7),
                ("ALIGN",         (0,0), (-1,-1), "LEFT"),
                ("BOX",           (0,0), (-1,-1), 1, GOLD),
                ("GRID",          (0,0), (-1,-1), 0.3, LIGHT_GOLD),
                ("TOPPADDING",    (0,0), (-1,-1), 3),
                ("BOTTOMPADDING", (0,0), (-1,-1), 3),
            ]
            for i in range(1, len(am_data)):
                bg = MID_GREEN if i % 2 == 0 else DARK_GREEN
                am_style_list.append(("BACKGROUND", (0,i), (-1,i), bg))
                am_style_list.append(("TEXTCOLOR",  (0,i), (-1,i), WHITE))
            am_table.setStyle(TableStyle(am_style_list))
            story.append(am_table)
            story.append(Spacer(1, 5*mm))

            # ── Section 3: Banking Details ────────────────────────────────
            story.append(Paragraph("Banking & Debit Order Details", sub_section_style))
            bk_headers = ["Mem No", "Surname", "Account Holder",
                          "Bank", "Account No", "Branch", "Branch Code", "Debit Date"]
            bk_data = [bk_headers]
            for m in members:
                bk_data.append([
                    m[0] or "", m[2] or "", m[20] or "",
                    m[21] or "", m[22] or "", m[23] or "",
                    m[24] or "", m[25] or ""
                ])
            if len(bk_data) == 1:
                bk_data.append(["No members"] + ["-"] * 7)

            bk_col_w = [22*mm, 25*mm, 30*mm, 25*mm, 28*mm, 22*mm, 20*mm, 18*mm]
            bk_table = Table(bk_data, colWidths=bk_col_w, repeatRows=1)
            bk_style_list = [
                ("BACKGROUND",    (0,0), (-1,0), DARK_GREEN),
                ("TEXTCOLOR",     (0,0), (-1,0), GOLD),
                ("FONTNAME",      (0,0), (-1,0), "Helvetica-Bold"),
                ("FONTSIZE",      (0,0), (-1,-1), 7),
                ("ALIGN",         (0,0), (-1,-1), "LEFT"),
                ("BOX",           (0,0), (-1,-1), 1, GOLD),
                ("GRID",          (0,0), (-1,-1), 0.3, LIGHT_GOLD),
                ("TOPPADDING",    (0,0), (-1,-1), 3),
                ("BOTTOMPADDING", (0,0), (-1,-1), 3),
            ]
            for i in range(1, len(bk_data)):
                bg = MID_GREEN if i % 2 == 0 else DARK_GREEN
                bk_style_list.append(("BACKGROUND", (0,i), (-1,i), bg))
                bk_style_list.append(("TEXTCOLOR",  (0,i), (-1,i), WHITE))
            bk_table.setStyle(TableStyle(bk_style_list))
            story.append(bk_table)

            # ── Footer ───────────────────────────────────────────────────
            story.append(Spacer(1, 8*mm))
            story.append(HRFlowable(width="100%", thickness=1, color=GOLD))
            story.append(Paragraph(
                "Labour and Civic Organization (LACO)  |  lacosouthafrica@outlook.com  "
                "|  071 526 9250 / 064 776 8609  |  403 Delta Towers, Durban 4000",
                sub_style))

            # Build PDF
            doc.build(story)

            messagebox.showinfo("PDF Ready", f"Report saved to: {filepath}. Opening for printing...", parent=self)

            # Open for printing
            if platform.system() == "Darwin":
                subprocess.call(["open", filepath])
            elif platform.system() == "Windows":
                os.startfile(filepath)
            else:
                subprocess.call(["xdg-open", filepath])

        except Exception as e:
            messagebox.showerror("PDF Error", f"Could not generate PDF: {str(e)}", parent=self)


# ============================================================================
# MEMBER FORM WINDOW
# ============================================================================

class MemberForm(tk.Toplevel):
    def __init__(self, parent, app, member_id=None):
        super().__init__(parent)
        self.app = app
        self.member_id = member_id
        self.title("Edit Member" if member_id else "Add New Member")
        self.geometry("800x760")
        self.configure(bg=COLORS["bg"])
        self.resizable(True, True)
        self.grab_set()
        self.pdf_path = tk.StringVar()
        self.vars = {}
        self._build()
        if member_id:
            self._load_data()

    def _build(self):
        # Fixed header at top
        hdr = tk.Frame(self, bg=COLORS["header_bg"], pady=12)
        hdr.pack(fill="x", side="top")
        title = "Edit Member" if self.member_id else "Add New Member"
        tk.Label(hdr, text=title, font=("Arial", 15, "bold"),
                 fg="black", bg=COLORS["header_bg"]).pack()

        # !! Fixed Save bar pinned to BOTTOM - packed BEFORE canvas !!
        save_bar = tk.Frame(self, bg=COLORS["panel"], pady=10)
        save_bar.pack(fill="x", side="bottom")
        styled_button(save_bar, "SAVE MEMBER", self._save,
                      color="#2d7a2d", width=20).pack(side="left", padx=15)
        styled_button(save_bar, "Cancel", self.destroy,
                      color="#cc2200", width=12).pack(side="left", padx=5)
        tk.Label(save_bar, text="* Surname and First Name are required",
                 font=("Arial", 9), fg=COLORS["gold"],
                 bg=COLORS["panel"]).pack(side="right", padx=15)

        # Scrollable area fills the middle
        container = tk.Frame(self, bg=COLORS["bg"])
        container.pack(fill="both", expand=True, side="top")

        canvas = tk.Canvas(container, bg=COLORS["bg"], highlightthickness=0)
        vsb = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
        sf = tk.Frame(canvas, bg=COLORS["bg"])
        sf.bind("<Configure>",
                lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=sf, anchor="nw")
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        def scroll(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        canvas.bind_all("<MouseWheel>", scroll)

        # ── Build form fields ─────────────────────────────────────────────
        def section(text):
            tk.Label(sf, text=f"  {text}", font=("Arial", 11, "bold"),
                     fg=COLORS["gold"], bg=COLORS["card"],
                     anchor="w", pady=6).pack(fill="x", padx=15, pady=(18,4))

        def field(label, key, kind="entry", opts=None):
            rf = tk.Frame(sf, bg=COLORS["bg"])
            rf.pack(fill="x", padx=20, pady=3)
            tk.Label(rf, text=label, font=("Arial", 9), fg=COLORS["subtext"],
                     bg=COLORS["bg"], width=26, anchor="w").pack(side="left")
            var = tk.StringVar()
            self.vars[key] = var
            if kind == "combo":
                w = styled_combo(rf, opts or [], textvariable=var)
            elif kind == "text":
                w = tk.Text(rf, width=34, height=3, font=("Arial", 10),
                            bg=COLORS["entry_bg"], fg=COLORS["gold"],
                            insertbackground="black", relief="flat", bd=4)
                self.vars[key] = w   # store widget directly
            else:
                w = styled_entry(rf, textvariable=var)
            w.pack(side="left", padx=4)

        # Section A
        section("SECTION A: Personal Information")

        row1 = tk.Frame(sf, bg=COLORS["bg"])
        row1.pack(fill="x", padx=20, pady=3)
        tk.Label(row1, text="Title", font=("Arial", 9), fg=COLORS["subtext"],
                 bg=COLORS["bg"], width=26, anchor="w").pack(side="left")
        self.vars["title"] = tk.StringVar()
        styled_combo(row1, TITLES, textvariable=self.vars["title"], width=10).pack(side="left", padx=4)
        tk.Label(row1, text="  Gender", font=("Arial", 9),
                 fg=COLORS["subtext"], bg=COLORS["bg"]).pack(side="left")
        self.vars["gender"] = tk.StringVar()
        styled_combo(row1, ["Male","Female","Other"],
                     textvariable=self.vars["gender"], width=10).pack(side="left", padx=4)

        field("Surname *", "surname")
        field("First Name(s) *", "firstnames")

        row2 = tk.Frame(sf, bg=COLORS["bg"])
        row2.pack(fill="x", padx=20, pady=3)
        tk.Label(row2, text="Initials", font=("Arial", 9), fg=COLORS["subtext"],
                 bg=COLORS["bg"], width=26, anchor="w").pack(side="left")
        self.vars["initials"] = tk.StringVar()
        styled_entry(row2, width=8, textvariable=self.vars["initials"]).pack(side="left", padx=4)
        tk.Label(row2, text="  DOB (YYYY-MM-DD)", font=("Arial", 9),
                 fg=COLORS["subtext"], bg=COLORS["bg"]).pack(side="left")
        self.vars["date_of_birth"] = tk.StringVar()
        styled_entry(row2, width=14, textvariable=self.vars["date_of_birth"]).pack(side="left", padx=4)

        field("ID Number", "id_number")

        # Section B
        section("SECTION B: Contact Details")
        field("Postal Address", "postal_address", "text")
        field("Postal Code", "postal_code")

        row3 = tk.Frame(sf, bg=COLORS["bg"])
        row3.pack(fill="x", padx=20, pady=3)
        tk.Label(row3, text="Tel Number", font=("Arial", 9), fg=COLORS["subtext"],
                 bg=COLORS["bg"], width=26, anchor="w").pack(side="left")
        self.vars["tel_number"] = tk.StringVar()
        styled_entry(row3, width=15, textvariable=self.vars["tel_number"]).pack(side="left", padx=4)
        tk.Label(row3, text="  Cell No", font=("Arial", 9),
                 fg=COLORS["subtext"], bg=COLORS["bg"]).pack(side="left")
        self.vars["cell_number"] = tk.StringVar()
        styled_entry(row3, width=15, textvariable=self.vars["cell_number"]).pack(side="left", padx=4)

        field("Email", "email")
        field("Province", "province", "combo", PROVINCES)

        row4 = tk.Frame(sf, bg=COLORS["bg"])
        row4.pack(fill="x", padx=20, pady=3)
        tk.Label(row4, text="Municipality", font=("Arial", 9), fg=COLORS["subtext"],
                 bg=COLORS["bg"], width=26, anchor="w").pack(side="left")
        self.vars["municipality"] = tk.StringVar()
        styled_entry(row4, width=20, textvariable=self.vars["municipality"]).pack(side="left", padx=4)
        tk.Label(row4, text="  Ward", font=("Arial", 9),
                 fg=COLORS["subtext"], bg=COLORS["bg"]).pack(side="left")
        self.vars["ward"] = tk.StringVar()
        styled_entry(row4, width=8, textvariable=self.vars["ward"]).pack(side="left", padx=4)

        # Section C
        section("SECTION C: Membership & Banking")
        field("Membership Category", "membership_category", "combo", CATEGORIES)

        row5 = tk.Frame(sf, bg=COLORS["bg"])
        row5.pack(fill="x", padx=20, pady=3)
        tk.Label(row5, text="Monthly Amount (R)", font=("Arial", 9), fg=COLORS["subtext"],
                 bg=COLORS["bg"], width=26, anchor="w").pack(side="left")
        self.vars["monthly_amount"] = tk.StringVar(value="50")
        styled_entry(row5, width=10, textvariable=self.vars["monthly_amount"]).pack(side="left", padx=4)
        tk.Label(row5, text="  Status", font=("Arial", 9),
                 fg=COLORS["subtext"], bg=COLORS["bg"]).pack(side="left")
        self.vars["status"] = tk.StringVar(value="Pending")
        styled_combo(row5, STATUSES, textvariable=self.vars["status"], width=12).pack(side="left", padx=4)

        field("Date Joined (YYYY-MM-DD)", "date_joined")
        field("Account Holder Name", "bank_account_holder")
        field("Bank Name", "bank_name")

        row6 = tk.Frame(sf, bg=COLORS["bg"])
        row6.pack(fill="x", padx=20, pady=3)
        tk.Label(row6, text="Account Number", font=("Arial", 9), fg=COLORS["subtext"],
                 bg=COLORS["bg"], width=26, anchor="w").pack(side="left")
        self.vars["account_number"] = tk.StringVar()
        styled_entry(row6, width=18, textvariable=self.vars["account_number"]).pack(side="left", padx=4)
        tk.Label(row6, text="  Branch Code", font=("Arial", 9),
                 fg=COLORS["subtext"], bg=COLORS["bg"]).pack(side="left")
        self.vars["branch_code"] = tk.StringVar()
        styled_entry(row6, width=10, textvariable=self.vars["branch_code"]).pack(side="left", padx=4)

        field("Branch Name", "branch_name")
        field("Debit Order Date", "debit_order_date")

        # PDF
        section("Membership Form PDF")
        pdf_row = tk.Frame(sf, bg=COLORS["bg"])
        pdf_row.pack(fill="x", padx=20, pady=4)
        tk.Label(pdf_row, text="Uploaded PDF", font=("Arial", 9), fg=COLORS["subtext"],
                 bg=COLORS["bg"], width=26, anchor="w").pack(side="left")
        tk.Label(pdf_row, textvariable=self.pdf_path, font=("Arial", 9),
                 fg=COLORS["gold"], bg=COLORS["bg"], wraplength=320,
                 anchor="w").pack(side="left", padx=4)
        pdf_btn = tk.Frame(sf, bg=COLORS["bg"])
        pdf_btn.pack(fill="x", padx=20, pady=3)
        styled_button(pdf_btn, "Upload PDF", self._browse_pdf,
                      color="#c8a400", width=14).pack(side="left")

        # Notes
        section("Notes")
        notes_wrap = tk.Frame(sf, bg=COLORS["bg"])
        notes_wrap.pack(fill="x", padx=20, pady=(4, 24))
        self.notes_text = tk.Text(notes_wrap, width=64, height=3,
                                  font=("Arial", 10), bg=COLORS["entry_bg"],
                                  fg="#1a2f36", insertbackground="black",
                                  relief="flat", bd=4)
        self.notes_text.pack()

    def _browse_pdf(self):
        path = filedialog.askopenfilename(
            title="Select Membership Form PDF",
            filetypes=[("PDF", "*.pdf"), ("All", "*.*")])
        if path:
            self.pdf_path.set(path)

    def _get(self, key):
        w = self.vars.get(key)
        if w is None:
            return ""
        if isinstance(w, tk.Text):
            return w.get("1.0", tk.END).strip()
        return w.get().strip()

    def _load_data(self):
        data = get_member_by_id(self.member_id)
        if not data:
            return
        for key, var in self.vars.items():
            val = str(data.get(key, "") or "")
            if isinstance(var, tk.Text):
                var.delete("1.0", tk.END)
                var.insert("1.0", val)
            else:
                var.set(val)
        self.pdf_path.set(data.get("form_pdf_path", "") or "")
        self.notes_text.delete("1.0", tk.END)
        self.notes_text.insert("1.0", data.get("notes", "") or "")

    def _save(self):
        if not self._get("surname") or not self._get("firstnames"):
            messagebox.showerror("Required", "Surname and First Name are required.", parent=self)
            return
        try:
            amount = float(self._get("monthly_amount") or 50)
        except ValueError:
            amount = 50.0

        data = {
            "title":               self._get("title"),
            "surname":             self._get("surname"),
            "firstnames":          self._get("firstnames"),
            "initials":            self._get("initials"),
            "id_number":           self._get("id_number"),
            "date_of_birth":       self._get("date_of_birth"),
            "gender":              self._get("gender"),
            "postal_address":      self._get("postal_address"),
            "postal_code":         self._get("postal_code"),
            "tel_number":          self._get("tel_number"),
            "cell_number":         self._get("cell_number"),
            "email":               self._get("email"),
            "province":            self._get("province"),
            "ward":                self._get("ward"),
            "municipality":        self._get("municipality"),
            "membership_category": self._get("membership_category"),
            "monthly_amount":      amount,
            "bank_account_holder": self._get("bank_account_holder"),
            "bank_name":           self._get("bank_name"),
            "account_number":      self._get("account_number"),
            "branch_name":         self._get("branch_name"),
            "branch_code":         self._get("branch_code"),
            "debit_order_date":    self._get("debit_order_date"),
            "status":              self._get("status"),
            "date_joined":         self._get("date_joined") or datetime.now().strftime("%Y-%m-%d"),
            "form_pdf_path":       self.pdf_path.get(),
            "notes":               self.notes_text.get("1.0", tk.END).strip(),
        }

        if self.member_id:
            ok, msg = update_member(self.member_id, data)
        else:
            ok, msg, _, mem_no = add_member(data)
            if ok:
                msg = f"Member saved successfully!\nMembership No: {mem_no}"

        if ok:
            messagebox.showinfo("Saved", msg, parent=self)
            self.app.refresh()
            self.destroy()
        else:
            messagebox.showerror("Error", msg, parent=self)


# ============================================================================
# MAIN APP
# ============================================================================

class LacoApp:
    def __init__(self, root):
        self.root = root
        self.root.title("LACO Membership Management System")
        self.root.geometry("1200x720")
        self.root.minsize(900, 600)
        self.root.configure(bg=COLORS["bg"])
        create_tables()
        self._build_ui()
        self.refresh()

    def _build_ui(self):
        # Header
        header = tk.Frame(self.root, bg=COLORS["header_bg"], pady=8)
        header.pack(fill="x")
        # Logo
        try:
            logo_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "laco.jpg")
            if os.path.exists(logo_path):
                img = Image.open(logo_path).resize((70, 70), Image.Resampling.LANCZOS)
                self._logo = ImageTk.PhotoImage(img)
                tk.Label(header, image=self._logo, bg=COLORS["header_bg"]).pack(side="left", padx=10, pady=4)
        except Exception:
            pass
        title_frame = tk.Frame(header, bg=COLORS["header_bg"])
        title_frame.pack(side="left", padx=10)
        tk.Label(title_frame, text="LABOUR AND CIVIC ORGANIZATION",
                 font=("Arial", 17, "bold"), fg=COLORS["gold"],
                 bg=COLORS["header_bg"]).pack(anchor="w")
        tk.Label(title_frame, text="Membership Management System  v1.1",
                 font=("Arial", 10), fg=COLORS["subtext"],
                 bg=COLORS["header_bg"]).pack(anchor="w")

        # Stats bar
        stats_frame = tk.Frame(self.root, bg=COLORS["panel"], pady=8)
        stats_frame.pack(fill="x")
        self.stat_labels = {}
        for key, label, color in [
            ("total",           "Total Members",  COLORS["gold"]),
            ("active",          "Active",         COLORS["green"]),
            ("pending",         "Pending",        "#c8a400"),
            ("monthly_revenue", "Monthly Revenue",COLORS["accent"]),
        ]:
            card = tk.Frame(stats_frame, bg=COLORS["card"], padx=18, pady=6)
            card.pack(side="left", padx=10, pady=2)
            tk.Label(card, text=label, font=("Arial", 8),
                     fg=COLORS["subtext"], bg=COLORS["card"]).pack()
            lbl = tk.Label(card, text="--", font=("Arial", 14, "bold"),
                           fg=color, bg=COLORS["card"])
            lbl.pack()
            self.stat_labels[key] = lbl

        # Toolbar
        toolbar = tk.Frame(self.root, bg=COLORS["bg"], pady=8)
        toolbar.pack(fill="x", padx=15)
        styled_button(toolbar, "Add Member",  self._add_member,  color="#2d7a2d", width=13).pack(side="left", padx=4)
        styled_button(toolbar, "Edit",         self._edit_member, color="#c8a400", width=8).pack(side="left", padx=4)
        styled_button(toolbar, "View PDF",     self._open_pdf,    color="#c8a400", width=9).pack(side="left", padx=4)
        styled_button(toolbar, "Delete",       self._delete,      color="#cc2200", width=8).pack(side="left", padx=4)
        styled_button(toolbar, "Reports",      self._reports,     color="#c8a400", width=10).pack(side="left", padx=4)

        # Search
        sf = tk.Frame(toolbar, bg=COLORS["entry_bg"], padx=8, pady=4)
        sf.pack(side="right", padx=4)
        tk.Label(sf, text="Search:", bg=COLORS["entry_bg"],
                 fg=COLORS["subtext"], font=("Arial", 9)).pack(side="left")
        self.search_var = tk.StringVar()
        self.search_var.trace("w", self._live_search)
        tk.Entry(sf, textvariable=self.search_var, font=("Arial", 11),
                 bg=COLORS["entry_bg"], fg=COLORS["gold"],
                 insertbackground="black", relief="flat", width=28).pack(side="left", padx=4)
        tk.Button(sf, text="x", command=self._clear_search,
                  bg=COLORS["entry_bg"], fg=COLORS["subtext"],
                  relief="flat", cursor="hand2").pack(side="left")

        # Member list treeview
        tree_frame = tk.Frame(self.root, bg=COLORS["bg"])
        tree_frame.pack(fill="both", expand=True, padx=15, pady=(0,5))

        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Treeview",
                        background=COLORS["panel"], foreground=COLORS["gold"],
                        fieldbackground=COLORS["panel"], rowheight=26, font=("Arial", 10))
        style.configure("Treeview.Heading",
                        background=COLORS["card"], foreground=COLORS["accent"],
                        font=("Arial", 10, "bold"), relief="flat")
        style.map("Treeview",
                  background=[("selected", COLORS["red"])],
                  foreground=[("selected", COLORS["gold"])])

        cols = ("Mem No","Title","Surname","First Name","ID Number",
                "Cell","Email","Province","Category","Status","Joined")
        self.tree = ttk.Treeview(tree_frame, columns=cols, show="headings", selectmode="browse")
        widths    = [120, 50, 130, 130, 120, 110, 160, 100, 130, 80, 90]
        for col, w in zip(cols, widths):
            self.tree.heading(col, text=col, command=lambda c=col: self._sort(c))
            self.tree.column(col, width=w, anchor="w")

        vsb = ttk.Scrollbar(tree_frame, orient="vertical",   command=self.tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")
        self.tree.pack(fill="both", expand=True)
        self.tree.bind("<Double-1>", lambda e: self._edit_member())

        self.tree.tag_configure("Active",    foreground="#4dcc4d")
        self.tree.tag_configure("Pending",   foreground="#f0c800")
        self.tree.tag_configure("Inactive",  foreground="#999999")
        self.tree.tag_configure("Suspended", foreground="#ff4422")

        # Status bar
        self.status_bar = tk.Label(self.root, text="Ready", font=("Arial", 9),
                                   fg=COLORS["gold"], bg=COLORS["panel"],
                                   anchor="w", padx=10)
        self.status_bar.pack(fill="x", side="bottom")

    def refresh(self, data=None):
        for item in self.tree.get_children():
            self.tree.delete(item)
        rows = data if data is not None else get_all_members()
        for r in rows:
            status = r.get("status", "Pending")
            self.tree.insert("", "end", iid=str(r["id"]), tags=(status,),
                             values=(r.get("membership_number",""), r.get("title",""),
                                     r.get("surname",""), r.get("firstnames",""),
                                     r.get("id_number",""), r.get("cell_number",""),
                                     r.get("email",""), r.get("province",""),
                                     r.get("membership_category",""), status,
                                     r.get("date_joined","")))
        self.status_bar.config(text=f"Showing {len(rows)} member(s)")
        self._update_stats()

    def _update_stats(self):
        s = get_stats()
        self.stat_labels["total"].config(text=str(s["total"]))
        self.stat_labels["active"].config(text=str(s["active"]))
        self.stat_labels["pending"].config(text=str(s["pending"]))
        self.stat_labels["monthly_revenue"].config(text=f"R {s['monthly_revenue']:,.0f}")

    def _live_search(self, *args):
        term = self.search_var.get().strip()
        self.refresh(search_members(term) if term else None)

    def _clear_search(self):
        self.search_var.set("")
        self.refresh()

    def _sort(self, col):
        items = [(self.tree.set(k, col), k) for k in self.tree.get_children("")]
        items.sort()
        for i, (_, k) in enumerate(items):
            self.tree.move(k, "", i)

    def _selected_id(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning("Select Member", "Please select a member first.")
            return None
        return int(sel[0])

    def _add_member(self):
        MemberForm(self.root, self)

    def _edit_member(self):
        mid = self._selected_id()
        if mid:
            MemberForm(self.root, self, member_id=mid)

    def _reports(self):
        ReportsWindow(self.root)

    def _delete(self):
        mid = self._selected_id()
        if not mid:
            return
        data = get_member_by_id(mid)
        name = f"{data.get('firstnames','')} {data.get('surname','')}"
        if messagebox.askyesno("Confirm Delete", f"Delete '{name}'?\nThis cannot be undone."):
            delete_member(mid)
            self.refresh()

    def _open_pdf(self):
        mid = self._selected_id()
        if not mid:
            return
        data = get_member_by_id(mid)
        pdf = data.get("form_pdf_path", "")
        if not pdf or not os.path.exists(pdf):
            messagebox.showinfo("No PDF", "No PDF uploaded for this member.")
            return
        try:
            if platform.system() == "Darwin":
                subprocess.call(["open", pdf])
            elif platform.system() == "Windows":
                os.startfile(pdf)
            else:
                subprocess.call(["xdg-open", pdf])
        except Exception as e:
            messagebox.showerror("Error", str(e))


# ============================================================================
# ENTRY POINT
# ============================================================================

def main():
    root = tk.Tk()
    LacoApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()
